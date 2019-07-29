# -*- coding: utf-8 -*-

import os
import openpyxl
import re
from openpyxl.utils import get_column_letter

# Path to directory with working files
path = 'c:\Users\working\PycharmProjects\parse_xml\\'

# name of the file xlsx DEFAULT(source)
file_xls = 'vygruzka.xlsx'

# name of the new file xlsx with completed results
file_finish = 'itog.xlsx'

# open file xlsx DEFAULT(source)
wb = openpyxl.load_workbook(path + file_xls)

# choose sheet in slsx file
sheet = wb['Report Utilization']


# number of column basic first day input traffic utilization
input_traffic_column = 0
# number of column basic first day output traffic utilization
output_traffic_column = 1

for file in os.listdir(path):
    if 'cacti' in file:

        # open file csv
        my_file = open(path + file)
        print('*  '*3 + 'Parsing file: ' + file + '  *'*3)

        # looking for date in csv file
        date = re.findall('.*End:.* (\d*-\d*-\d*)', my_file.read(), flags=8)[0]

        # going to the first string of file for next iterations
        my_file.seek(0)

        # next column input traffic utilization
        input_traffic_column += 2

        # next column output traffic utilization
        output_traffic_column += 2

        for line in my_file:
            # looking for string starting with 'PO' - start DNS name of device
            if 'PO' in line.split(';')[0]:

                # write date of file to ROW 16
                sheet[get_column_letter(input_traffic_column)+ str(16)].value = date

                # splitting first item from string - device name and port in csv file
                interface = line.split(';')[0].split('"')[1]

                # splitting input traffic usage from csv file
                input_traffic = line.split(';')[8].split('"')[1]

                # splitting output traffic usage from csv file
                output_traffic = line.split(';')[10].split('"')[1]

                # looking for match string with device-port csv file with cell of xlsx file
                for cellObj in sheet['a17':'a425']:
                    for cell in cellObj:
                        if cell.value == interface:

                            line_num = cell.row

                            # get column_letter1 for input_traffic
                            column_letter1 = get_column_letter(input_traffic_column)

                            # get column_letter1 for output_traffic
                            column_letter2 = get_column_letter(output_traffic_column)

                            # write data from input traffic csv file to xlsx file
                            sheet[column_letter1 + str(line_num)].value = input_traffic

                            # write data from ouput traffic csv file to xlsx file
                            sheet[column_letter2 + str(line_num)].value = output_traffic

                            break
# Write data to new file 'itog.xlsx'
wb.save(path + file_finish)
